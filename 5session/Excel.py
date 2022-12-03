import openpyxl

wb_sum = openpyxl.load_workbook("出社在宅集計表_まとめ.xlsx")




month_list = wb_sum.sheetnames


department_list = ["人事部","企画部","営業部"]

for department in department_list:
    wb_department = openpyxl.load_workbook("出社在宅集計表_{}.xlsx".format(department))    
    
    
    for month in month_list:
        ws_sum = wb_sum[month]        
        ws_kakubu = wb_department[month]
        
        
        for j in range(1, ws_sum.max_row + 1) :
            if ws_sum.cell(row=j, column=1).value==department:
                syussya_row = j
                zaitaku_row = j+1

        for i in range(ws_kakubu.max_column -1):
            ws_sum.cell(row=syussya_row,column=3+i).value = ws_kakubu.cell(row=2,column=2+i).value
            ws_sum.cell(row=zaitaku_row,column=3+i).value = ws_kakubu.cell(row=3,column=2+i).value


wb_sum.save("test_sum.xlsx")
